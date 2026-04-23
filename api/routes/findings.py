import json
import logging
import csv
from math import ceil
from io import BytesIO, StringIO
from datetime import datetime, timezone, timedelta

from fastapi import APIRouter, Depends, Query
from fastapi.responses import Response
from pydantic import BaseModel
from sqlalchemy import select, func
from sqlalchemy.ext.asyncio import AsyncSession

from core.database import get_db
from core.models import Finding

logger = logging.getLogger(__name__)

BRT = timezone(timedelta(hours=-3))

router = APIRouter(tags=["findings"])


class StatusUpdate(BaseModel):
    status: str


def _serialize_finding(f: Finding) -> dict:
    try:
        snippets_data = json.loads(f.snippets) if f.snippets else []
    except (json.JSONDecodeError, TypeError):
        snippets_data = []

    try:
        terms_data = json.loads(f.sensitive_terms) if f.sensitive_terms else []
    except (json.JSONDecodeError, TypeError):
        terms_data = []

    try:
        meta_data = json.loads(f.doc_metadata) if f.doc_metadata else {}
    except (json.JSONDecodeError, TypeError):
        meta_data = {}

    return {
        "id": f.id,
        "scan_id": f.scan_id,
        "discovered_at": f.discovered_at.isoformat() if f.discovered_at else None,
        "source_platform": f.source_platform,
        "url": f.url,
        "title": f.title,
        "file_type": f.file_type,
        "risk_level": f.risk_level,
        "risk_score": f.risk_score,
        "category": f.category,
        "entity_matched": f.entity_matched,
        "cpf_count": f.cpf_count,
        "cnpj_count": f.cnpj_count,
        "financial_count": f.financial_count,
        "sensitive_terms": terms_data,
        "snippets": snippets_data,
        "author": f.author or "",
        "publisher": f.publisher or "",
        "country": f.country or "INT",
        "language": f.language or "unknown",
        "doc_metadata": meta_data,
        "resolution_status": f.resolution_status,
        "analyst_notes": f.analyst_notes,
    }


STATUS_LABELS = {
    "pending": "Pendente", "investigating": "Investigando",
    "false_positive": "Falso Positivo", "auto_false_positive": "FP Automatico",
    "resolved": "Resolvido", "notified": "Notificado",
}
RISK_LABELS = {
    "critical": "CRITICO", "high": "ALTO",
    "medium": "MEDIO", "low": "BAIXO",
}

EXPORT_HEADERS = [
    "ID", "Data", "Score", "Risco", "Status", "Pais", "Idioma",
    "Entidade", "Plataforma", "Titulo", "Tipo", "Autor",
    "CPFs", "CNPJs", "Financeiros", "Categoria", "URL", "Notas",
]


def _build_filter_query(risk_level, status, category, country, language):
    base = select(Finding).where(Finding.is_deleted == False)
    if risk_level:
        base = base.where(Finding.risk_level == risk_level)
    if status:
        base = base.where(Finding.resolution_status == status)
    if category:
        base = base.where(Finding.category == category)
    if country:
        base = base.where(Finding.country == country)
    if language:
        base = base.where(Finding.language == language)
    return base.order_by(Finding.risk_score.desc(), Finding.id.desc())


def _finding_to_row(f: Finding) -> list:
    date_str = f.discovered_at.strftime("%d/%m/%Y %H:%M") if f.discovered_at else ""
    return [
        f.id, date_str, f.risk_score,
        RISK_LABELS.get(f.risk_level, f.risk_level or ""),
        STATUS_LABELS.get(f.resolution_status, f.resolution_status or ""),
        f.country or "INT", f.language or "",
        f.entity_matched or "", f.source_platform or "", f.title or "",
        (f.file_type or "").upper(), f.author or "",
        f.cpf_count or 0, f.cnpj_count or 0, f.financial_count or 0,
        f.category or "", f.url or "", f.analyst_notes or "",
    ]


@router.get("/export/xlsx")
async def export_xlsx(
    risk_level: str = Query(""),
    status: str = Query(""),
    category: str = Query(""),
    country: str = Query(""),
    language: str = Query(""),
    db: AsyncSession = Depends(get_db),
):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    query = _build_filter_query(risk_level, status, category, country, language)
    rows = (await db.execute(query)).scalars().all()
    logger.info("XLSX export: %d findings", len(rows))

    wb = Workbook()
    ws = wb.active
    ws.title = "OSINT Findings"

    hdr_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    hdr_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    hdr_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )

    risk_fills = {
        "critical": PatternFill(start_color="FFE0E5", end_color="FFE0E5", fill_type="solid"),
        "high": PatternFill(start_color="FFF0DB", end_color="FFF0DB", fill_type="solid"),
        "medium": PatternFill(start_color="FFFBE6", end_color="FFFBE6", fill_type="solid"),
        "low": PatternFill(start_color="E0FFF5", end_color="E0FFF5", fill_type="solid"),
    }

    for col_idx, header in enumerate(EXPORT_HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = hdr_align
        cell.border = thin_border

    for row_idx, f in enumerate(rows, 2):
        values = _finding_to_row(f)
        row_fill = risk_fills.get(f.risk_level)
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            cell.font = Font(name="Calibri", size=10)
            if row_fill:
                cell.fill = row_fill

    col_widths = [6, 16, 8, 10, 14, 8, 8, 25, 14, 35, 8, 18, 6, 6, 10, 14, 50, 30]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    buf = BytesIO()
    wb.save(buf)
    length = buf.tell()
    buf.seek(0)

    filename = f"osint_findings_{datetime.now(BRT).strftime('%Y%m%d_%H%M')}.xlsx"

    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Content-Length": str(length),
        },
    )


@router.get("/export/csv")
async def export_csv(
    risk_level: str = Query(""),
    status: str = Query(""),
    category: str = Query(""),
    country: str = Query(""),
    language: str = Query(""),
    db: AsyncSession = Depends(get_db),
):
    query = _build_filter_query(risk_level, status, category, country, language)
    rows = (await db.execute(query)).scalars().all()
    logger.info("CSV export: %d findings", len(rows))

    output = StringIO()
    writer = csv.writer(output, delimiter=";", quoting=csv.QUOTE_MINIMAL)
    writer.writerow(EXPORT_HEADERS)
    for f in rows:
        writer.writerow(_finding_to_row(f))

    content = output.getvalue().encode("utf-8-sig")
    filename = f"osint_findings_{datetime.now(BRT).strftime('%Y%m%d_%H%M')}.csv"

    return Response(
        content=content,
        media_type="text/csv; charset=utf-8",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Content-Length": str(len(content)),
        },
    )


@router.get("/findings")
async def list_findings(
    page: int = Query(1, ge=1),
    per_page: int = Query(20, ge=1, le=100),
    risk_level: str = Query(""),
    status: str = Query(""),
    category: str = Query(""),
    country: str = Query(""),
    language: str = Query(""),
    db: AsyncSession = Depends(get_db),
):
    base = select(Finding).where(Finding.is_deleted == False)
    if risk_level:
        base = base.where(Finding.risk_level == risk_level)
    if status:
        base = base.where(Finding.resolution_status == status)
    if category:
        base = base.where(Finding.category == category)
    if country:
        base = base.where(Finding.country == country)
    if language:
        base = base.where(Finding.language == language)

    count_q = select(func.count()).select_from(base.subquery())
    total = (await db.execute(count_q)).scalar() or 0
    pages = ceil(total / per_page) if total > 0 else 1

    query = base.order_by(Finding.risk_score.desc(), Finding.id.desc())
    query = query.offset((page - 1) * per_page).limit(per_page)
    rows = (await db.execute(query)).scalars().all()

    return {
        "total": total,
        "page": page,
        "per_page": per_page,
        "pages": pages,
        "findings": [_serialize_finding(f) for f in rows],
    }


@router.get("/findings/{finding_id}")
async def get_finding(
    finding_id: int,
    db: AsyncSession = Depends(get_db),
):
    from fastapi import HTTPException

    result = await db.execute(
        select(Finding).where(Finding.id == finding_id, Finding.is_deleted == False)
    )
    finding = result.scalar_one_or_none()
    if not finding:
        raise HTTPException(status_code=404, detail="Finding not found")
    return _serialize_finding(finding)


@router.patch("/findings/{finding_id}/status")
async def update_finding_status(
    finding_id: int,
    body: StatusUpdate,
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(select(Finding).where(Finding.id == finding_id))
    finding = result.scalar_one_or_none()

    if not finding:
        from fastapi import HTTPException
        raise HTTPException(status_code=404, detail="Finding not found")

    finding.resolution_status = body.status
    await db.commit()

    return {"id": finding.id, "status": finding.resolution_status}


@router.delete("/findings/{finding_id}")
async def soft_delete_finding(
    finding_id: int,
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(select(Finding).where(Finding.id == finding_id))
    finding = result.scalar_one_or_none()

    if not finding:
        from fastapi import HTTPException
        raise HTTPException(status_code=404, detail="Finding not found")

    finding.is_deleted = True
    await db.commit()

    return {"id": finding.id, "deleted": True}
