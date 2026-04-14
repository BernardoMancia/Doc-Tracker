import io
import logging
import re
from dataclasses import dataclass, field

import fitz
from docx import Document
from openpyxl import load_workbook

logger = logging.getLogger(__name__)

USER_AGENT_PATTERN = re.compile(r"Mozilla/|AppleWebKit/|Chrome/|Safari/|Gecko/|Firefox/|MSIE |Trident/", re.IGNORECASE)
TOOL_PATTERNS = [
    "microsoft", "libreoffice", "openoffice", "google docs", "adobe",
    "pdf-xchange", "nitro", "foxit", "pdflatex", "tex",
    "wkhtmltopdf", "prince", "reportlab", "itext", "cairo",
    "skia", "qt ", "cups", "ghostscript",
]


def _clean_metadata_field(value: str) -> str:
    if not value or not value.strip():
        return ""
    v = value.strip()
    if USER_AGENT_PATTERN.search(v):
        return ""
    v_lower = v.lower()
    for tool in TOOL_PATTERNS:
        if tool in v_lower:
            return ""
    if len(v) > 200:
        return ""
    return v


@dataclass
class ExtractionResult:
    text: str = ""
    metadata: dict = field(default_factory=dict)


class TextExtractor:

    def extract(self, data: io.BytesIO, file_type: str) -> ExtractionResult:
        data.seek(0)
        try:
            if file_type == "pdf":
                return self._extract_pdf(data)
            elif file_type == "docx":
                return self._extract_docx(data)
            elif file_type == "xlsx":
                return self._extract_xlsx(data)
            elif file_type in ("txt", "csv"):
                return self._extract_text(data)
            else:
                return self._extract_text(data)
        except Exception as e:
            logger.debug("Extraction failed for %s: %s", file_type, e)
            return ExtractionResult()

    def _extract_pdf(self, data: io.BytesIO) -> ExtractionResult:
        doc = fitz.open(stream=data.read(), filetype="pdf")
        pages = []
        for page in doc:
            pages.append(page.get_text())

        meta = doc.metadata or {}
        doc.close()

        author = _clean_metadata_field(meta.get("author", ""))
        creator = _clean_metadata_field(meta.get("creator", ""))
        producer = meta.get("producer", "")

        return ExtractionResult(
            text="\n".join(pages),
            metadata={
                "author": author,
                "creator": creator,
                "producer": producer,
                "title": meta.get("title", ""),
                "subject": meta.get("subject", ""),
                "creation_date": meta.get("creationDate", ""),
                "mod_date": meta.get("modDate", ""),
                "page_count": len(pages),
            },
        )

    def _extract_docx(self, data: io.BytesIO) -> ExtractionResult:
        doc = Document(data)
        paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]

        meta = {}
        if doc.core_properties:
            cp = doc.core_properties
            meta = {
                "author": _clean_metadata_field(cp.author or ""),
                "last_modified_by": _clean_metadata_field(cp.last_modified_by or ""),
                "title": cp.title or "",
                "subject": cp.subject or "",
                "created": str(cp.created) if cp.created else "",
                "modified": str(cp.modified) if cp.modified else "",
            }

        return ExtractionResult(text="\n".join(paragraphs), metadata=meta)

    def _extract_xlsx(self, data: io.BytesIO) -> ExtractionResult:
        wb = load_workbook(data, read_only=True, data_only=True)
        rows = []

        meta = {}
        if wb.properties:
            p = wb.properties
            meta = {
                "author": _clean_metadata_field(p.creator or ""),
                "last_modified_by": _clean_metadata_field(p.lastModifiedBy or ""),
                "title": p.title or "",
                "created": str(p.created) if p.created else "",
            }

        for sheet in wb.sheetnames:
            ws = wb[sheet]
            for row in ws.iter_rows(values_only=True):
                cells = [str(c) for c in row if c is not None]
                if cells:
                    rows.append(" | ".join(cells))
        wb.close()

        return ExtractionResult(text="\n".join(rows), metadata=meta)

    def _extract_text(self, data: io.BytesIO) -> ExtractionResult:
        raw = data.read()
        for encoding in ("utf-8", "latin-1", "cp1252"):
            try:
                return ExtractionResult(text=raw.decode(encoding))
            except (UnicodeDecodeError, Exception):
                continue
        return ExtractionResult()
